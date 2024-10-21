import fnmatch
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenPyXLImage
from openpyxl.styles import PatternFill
from PIL import Image as PILImage
from modules.db_utils import setup_ssh_client
import os
import re

def flatten_json(y):
    out = {}

    def flatten(x, name=''):
        if isinstance(x, dict):
            for a in x:
                flatten(x[a], name + a + '_')
        else:
            key = name[:-1]  # Remove the trailing underscore
            # Remove 'Properties_' prefix if it exists
            if key.startswith('Properties_'):
                key = key.replace('Properties_', '')
            out[key] = x

    flatten(y)
    return out

def retrieve_images(proxmox_auth, remote_path, local_path):

    try:
        # Set up SSH client and SFTP connection
        ssh_client = setup_ssh_client(proxmox_auth)
        sftp = ssh_client.open_sftp()

        # List files in the remote directory
        file_list = sftp.listdir(remote_path)
        #print(f"Files in remote directory: {file_list}")

        # Filter files based on a pattern (e.g., '*.jpg', or '*' for all files)
        matching_files = fnmatch.filter(file_list, '*')  # Modify pattern if needed
        downloaded_files = []
        #print(f"Matching files: {matching_files}")

        for remote_file in matching_files:
            # Construct full remote and local file paths
            remote_file_path = os.path.join(remote_path, remote_file)
            local_file_path = os.path.join(local_path, remote_file)

            # Download the file from remote to local
            sftp.get(remote_file_path, local_file_path)
            downloaded_files.append(local_file_path)
            #print(f"Downloaded: {remote_file_path} to {local_file_path}")
        
        # Close the SFTP connection
        sftp.close()
        return downloaded_files

    except Exception as e:
        print(f"Error downloading files: {e}")
        return []

def save_as_excel(data,proxmox_auth):

    #flatten the data
    flattened_data = [flatten_json(item) for item in data]

    #print("FLAT",flattened_data)
    wb = Workbook()
    ws = wb.active

    local_image_folder = "./downloaded_images"  # Folder to store downloaded images locally

    # Create folder if it doesn't exist
    os.makedirs(local_image_folder, exist_ok=True)

    # Define alternating colors: "forest green" and "pastel green"
    forest_green_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")  # light green
    pastel_green_fill = PatternFill(start_color="C1E1C1", end_color="C1E1C1", fill_type="solid")  # Pastel green

    # Adding headers
    headers = list(flattened_data[0].keys())  # Assumes all JSON entries have the same structure
    #print(headers)
    
    for idx, header in enumerate(headers):
        if header in ['Images_PackingSlip', 'Images_SerialNumber']:
                header = "                "
        if header == 'Images_Directory':
            header = 'images'
        ws.cell(row=1, column=idx + 1, value=header)
        # Adjust the width of the columns based on the header length
        column_letter = chr(65 + idx)  # Convert index to Excel column letter ('A', 'B', etc.)
        ws.column_dimensions[column_letter].width = max(15, len(header) + 2)  # Set column width

    image_count = {}

    for row_idx, entry in enumerate(flattened_data, start=2):  # Start from the second row
        # Choose the fill color based on the row number (alternating rows)
        fill_color = forest_green_fill if row_idx % 2 == 0 else pastel_green_fill
        
        # Insert non-image data
        for col_idx, (key, value) in enumerate(entry.items()):
           
            if key in ['Images_PackingSlip', 'Images_SerialNumber']:
                continue
            if not key.endswith('Images_Directory'):
                cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
                cell.fill = fill_color  # Apply the alternating fill color
            
            else:
                    # Extract image directory from the Images field   
                    image_directory = re.sub(r"smb@\d{1,3}(?:\.\d{1,3}){3}:", "", value)
                 
                    # Download or use the local path from `image_directory` (mock download)
                    downloaded_images = retrieve_images(proxmox_auth,image_directory,local_image_folder)

                    # Count the number of images inserted into this cell
                    image_count[row_idx] = len(downloaded_images)

                     # Set row height based on the image size (e.g., 100 pixels)
                    img_height = 100  # Height of the image after resizing
                    img_width = 100
                    ws.row_dimensions[row_idx].height = img_height+5.0 * 0.75  # Adjust row height (Excel's row height units)
                    
                    for img_idx, img_path in enumerate(downloaded_images, start=1):
                        if os.path.exists(img_path):  # Ensure the image file exists
                            img = PILImage.open(img_path)
                            img = img.resize((img_height, img_width), PILImage.ANTIALIAS)
                            img.save(img_path)  # Ensure it's saved before inserting
                            
                            excel_img = OpenPyXLImage(img_path)
                            #image_col = col_idx + 1  # Adjust based on your desired column
                            image_col = col_idx + img_idx
                         
                            image_row = row_idx  # Adjust row to position images closer
                          
                            # Insert image into the Excel sheet in the appropriate cell
                            excel_img.anchor = ws.cell(row=int(image_row), column=int(image_col)).coordinate  # Adjust column position for multiple images
                            
                            ws.add_image(excel_img)

                            # Set the width based on the number of images, adjusting as necessary
                            image_column_width = 20  # Base width for the image column
                            image_col_letter = get_column_letter(image_col)
                            # Update the image directory column width using the column letter
                            ws.column_dimensions[image_col_letter].width = image_column_width  # Update the image directory column width


    # Save the workbook
    wb.save("inventory.xlsx")
    print("Excel file created with images from SFTP.")

def pull_mongodb(client):
    db = client["local"]
    items = db['inventory'].find()

    inventory = []
    for item in items:
        inventory.append(item)
    
    return inventory

    

def pull_proxmox(proxmox_auth):
    host = proxmox_auth["host"]
    port = proxmox_auth["port"]
    user = proxmox_auth["user"]
    password = proxmox_auth["password"]

    inv_path = "/mnt/proxmox/images/inventory/"